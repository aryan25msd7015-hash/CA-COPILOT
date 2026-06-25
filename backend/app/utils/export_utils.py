"""Export utilities for Excel workbook generation."""
import io
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def build_reconciliation_excel(
    matched_rows: list[dict],
    unmatched_rows: list[dict],
    summary: dict | None = None,
) -> bytes:
    """Build a reconciliation workbook with summary, matched, and unmatched sheets."""
    wb = openpyxl.Workbook()
    summary = summary or {}

    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_sheet(ws_summary, ["Metric", "Value"], [
        ["Client", summary.get("client_name", "")],
        ["Period", summary.get("period", "")],
        ["Status", summary.get("status", "")],
        ["Total Purchase", summary.get("total_purchase", 0)],
        ["Total GSTR-2B", summary.get("total_gstr2b", 0)],
        ["Matched Count", summary.get("matched_count", 0)],
        ["Unmatched Count", summary.get("unmatched_count", 0)],
        ["Mismatch Value", summary.get("mismatch_value", 0)],
        ["Run At", str(summary.get("run_at", ""))],
        ["Completed At", str(summary.get("completed_at", ""))],
    ])

    ws_matched = wb.create_sheet("Matched")
    matched_headers = [
        "Invoice No",
        "Vendor Name",
        "Vendor GSTIN",
        "Amount",
        "Date",
        "Match Type",
        "Confidence %",
    ]
    _write_sheet(ws_matched, matched_headers, [
        [
            row.get("invoice_no", ""),
            row.get("vendor_name", ""),
            row.get("vendor_gstin", ""),
            row.get("amount", ""),
            str(row.get("date", "")),
            row.get("match_type", ""),
            row.get("confidence", ""),
        ]
        for row in matched_rows
    ])

    ws_unmatched = wb.create_sheet("Unmatched")
    unmatched_headers = ["Invoice No", "Vendor Name", "Vendor GSTIN", "Amount", "Date", "Source"]
    _write_sheet(ws_unmatched, unmatched_headers, [
        [
            row.get("invoice_no", ""),
            row.get("vendor_name", ""),
            row.get("vendor_gstin", ""),
            row.get("amount", ""),
            str(row.get("date", "")),
            row.get("source", ""),
        ]
        for row in unmatched_rows
    ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_sheet(ws, headers: list[str], rows: list[list[Any]]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(15, len(header) + 4)

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
